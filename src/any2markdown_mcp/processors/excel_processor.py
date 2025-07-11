"""
Excel文档处理器

基于pandas和openpyxl实现Excel转换功能，支持：
- Excel工作表转Markdown表格
- 多工作表处理
- 公式转换
- 图表处理
"""

import asyncio
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import re

import structlog
import pandas as pd
import openpyxl
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, ScatterChart
import io

from .base_processor import BaseProcessor

logger = structlog.get_logger(__name__)


class ExcelProcessor(BaseProcessor):
    """Excel文档处理器"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        
        # Excel特定配置
        self.max_rows_per_sheet = self._get_config_value("excel_max_rows_per_sheet", 10000)
        self.include_empty_cells = self._get_config_value("excel_include_empty_cells", False)
        self.convert_formulas = self._get_config_value("excel_convert_formulas", True)
        
        logger.info("ExcelProcessor initialized")
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的文件格式"""
        return ["xlsx", "xls"]
    
    async def convert(self, file_content: bytes, options: Dict[str, Any]) -> Dict[str, Any]:
        """
        转换Excel文档
        
        Args:
            file_content: Excel文件内容
            options: 转换选项
            
        Returns:
            转换结果
        """
        temp_files = []
        
        try:
            # 验证文件大小
            self.validate_file_size(file_content)
            
            # 保存临时Excel文件
            temp_excel_path = await self.save_temp_file(file_content, suffix=".xlsx")
            temp_files.append(str(temp_excel_path))
            
            logger.info("Starting Excel conversion", 
                       file_size=len(file_content),
                       temp_path=str(temp_excel_path))
            
            # 分析Excel结构
            excel_info = await self._analyze_excel_structure(temp_excel_path)
            
            # 提取图片（如果启用）
            images_info = []
            if options.get("extract_images", True):
                images_info = await self._extract_images(temp_excel_path)
                # 注意：不要将图片文件添加到temp_files中，因为它们需要被保留用于静态文件服务
            
            # 转换工作表
            markdown_content = await self._convert_worksheets(temp_excel_path, excel_info, options)
            
            # 嵌入图片URL到markdown中（如果有图片）
            if images_info:
                # 对于Excel，需要处理每个工作表的内容
                for sheet_data in markdown_content:
                    sheet_data["content"] = await self._embed_image_urls(sheet_data["content"], images_info)
            
            # 分页处理（按工作表分页）
            if options.get("paginate_output", True):
                pages = await self._paginate_by_sheets(markdown_content, excel_info)
            else:
                # 合并所有工作表内容
                combined_content = "\n\n".join([sheet["content"] for sheet in markdown_content])
                pages = [{"page": 1, "content": combined_content}]
            
            # 构建元数据
            metadata = {
                "source_type": "excel",
                "sheet_count": excel_info["sheet_count"],
                "sheet_names": excel_info["sheet_names"],
                "total_rows": excel_info["total_rows"],
                "total_columns": excel_info["total_columns"],
                "has_charts": excel_info["has_charts"],
                "images": images_info,
                "file_size": len(file_content),
                "options_used": options
            }
            
            # 格式化输出
            result = self.format_output(
                combined_content if not options.get("paginate_output", True) else pages,
                metadata,
                options.get("output_format", "markdown")
            )
            
            logger.info("Excel conversion completed successfully",
                       sheets_processed=excel_info["sheet_count"],
                       images_extracted=len(images_info))
            
            return result
            
        except Exception as e:
            logger.error("Excel conversion failed", error=str(e))
            raise
        finally:
            # 清理临时文件
            if temp_files:
                await self.cleanup_temp_files(temp_files)
    
    async def _analyze_excel_structure(self, excel_path: Path) -> Dict[str, Any]:
        """分析Excel文档结构"""
        try:
            # 使用openpyxl分析结构
            workbook = openpyxl.load_workbook(str(excel_path), read_only=True)
            
            info = {
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "total_rows": 0,
                "total_columns": 0,
                "has_charts": False,
                "has_formulas": False,
                "sheet_info": {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 获取有效数据范围
                if sheet.max_row and sheet.max_column:
                    sheet_rows = sheet.max_row
                    sheet_cols = sheet.max_column
                else:
                    sheet_rows = 0
                    sheet_cols = 0
                
                info["total_rows"] += sheet_rows
                info["total_columns"] = max(info["total_columns"], sheet_cols)
                
                # 工作表信息
                info["sheet_info"][sheet_name] = {
                    "rows": sheet_rows,
                    "columns": sheet_cols,
                    "has_data": sheet_rows > 0 and sheet_cols > 0
                }
            
            workbook.close()
            
            # 使用pandas检查公式和图表
            try:
                with pd.ExcelFile(str(excel_path)) as excel_file:
                    for sheet_name in excel_file.sheet_names:
                        # 读取少量数据检查是否有公式
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=10)
                        # 这里可以添加更详细的公式检测逻辑
            except Exception as e:
                logger.warning("Could not analyze Excel with pandas", error=str(e))
            
            logger.debug("Excel structure analyzed", info=info)
            return info
            
        except Exception as e:
            logger.error("Failed to analyze Excel structure", error=str(e))
            raise
    
    async def _convert_worksheets(self, excel_path: Path, excel_info: Dict[str, Any], options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """转换所有工作表"""
        worksheets_content = []
        
        try:
            # 指定要处理的工作表
            target_sheets = options.get("sheets", excel_info["sheet_names"])
            if target_sheets is None:
                target_sheets = excel_info["sheet_names"]
            if isinstance(target_sheets, str):
                target_sheets = [target_sheets]
            
            for sheet_name in target_sheets:
                if sheet_name not in excel_info["sheet_names"]:
                    logger.warning("Sheet not found", sheet_name=sheet_name)
                    continue
                
                logger.info("Converting sheet", sheet_name=sheet_name)
                
                try:
                    # 转换单个工作表
                    sheet_content = await self._convert_single_sheet(excel_path, sheet_name, options)
                    
                    worksheets_content.append({
                        "sheet_name": sheet_name,
                        "content": sheet_content
                    })
                    
                except Exception as e:
                    logger.error("Failed to convert sheet", 
                               sheet_name=sheet_name, error=str(e))
                    # 继续处理其他工作表
                    continue
            
            return worksheets_content
            
        except Exception as e:
            logger.error("Failed to convert worksheets", error=str(e))
            raise
    
    async def _convert_single_sheet(self, excel_path: Path, sheet_name: str, options: Dict[str, Any]) -> str:
        """转换单个工作表"""
        try:
            # 读取Excel数据
            df = pd.read_excel(
                str(excel_path), 
                sheet_name=sheet_name,
                na_filter=not self.include_empty_cells,
                keep_default_na=not self.include_empty_cells
            )
            
            # 限制行数
            if len(df) > self.max_rows_per_sheet:
                logger.warning("Sheet has too many rows, truncating", 
                             sheet_name=sheet_name,
                             original_rows=len(df),
                             max_rows=self.max_rows_per_sheet)
                df = df.head(self.max_rows_per_sheet)
            
            # 处理空值
            if not self.include_empty_cells:
                df = df.fillna("")
            
            # 转换为Markdown表格
            markdown_content = await self._dataframe_to_markdown(df, sheet_name, options)
            
            return markdown_content
            
        except Exception as e:
            logger.error("Failed to convert single sheet", 
                       sheet_name=sheet_name, error=str(e))
            raise
    
    async def _dataframe_to_markdown(self, df: pd.DataFrame, sheet_name: str, options: Dict[str, Any]) -> str:
        """将DataFrame转换为Markdown"""
        try:
            markdown_lines = []
            
            # 添加工作表标题
            markdown_lines.append(f"## {sheet_name}")
            markdown_lines.append("")
            
            # 检查是否有数据
            if df.empty:
                markdown_lines.append("*This sheet is empty.*")
                return "\n".join(markdown_lines)
            
            # 处理列名
            columns = []
            for col in df.columns:
                col_name = str(col).strip()
                if col_name.startswith("Unnamed:"):
                    col_name = f"Column {col_name.split(':')[1]}"
                columns.append(col_name)
            
            # 创建表格头部
            markdown_lines.append("| " + " | ".join(columns) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(columns)) + " |")
            
            # 添加数据行
            for index, row in df.iterrows():
                cells = []
                for col in df.columns:
                    cell_value = row[col]
                    
                    # 处理不同类型的值
                    if pd.isna(cell_value):
                        cell_str = ""
                    elif isinstance(cell_value, (int, float)):
                        if pd.isna(cell_value):
                            cell_str = ""
                        else:
                            cell_str = str(cell_value)
                    else:
                        cell_str = str(cell_value).strip()
                    
                    # 转义Markdown特殊字符
                    cell_str = cell_str.replace("|", "\\|").replace("\n", " ")
                    cells.append(cell_str)
                
                markdown_lines.append("| " + " | ".join(cells) + " |")
            
            # 添加统计信息
            markdown_lines.append("")
            markdown_lines.append(f"*Sheet contains {len(df)} rows and {len(df.columns)} columns.*")
            
            return "\n".join(markdown_lines)
            
        except Exception as e:
            logger.error("Failed to convert DataFrame to markdown", error=str(e))
            raise
    
    async def _paginate_by_sheets(self, worksheets_content: List[Dict[str, Any]], excel_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """按工作表分页"""
        pages = []
        
        for i, sheet_data in enumerate(worksheets_content):
            pages.append({
                "page": i + 1,
                "sheet_name": sheet_data["sheet_name"],
                "content": sheet_data["content"]
            })
        
        logger.info("Excel content paginated by sheets", total_pages=len(pages))
        return pages
    
    async def _extract_charts(self, excel_path: Path) -> List[Dict[str, Any]]:
        """提取Excel图表（可选功能）"""
        charts_info = []
        
        try:
            workbook = openpyxl.load_workbook(str(excel_path))
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 检查图表
                if hasattr(sheet, '_charts') and sheet._charts:
                    for i, chart in enumerate(sheet._charts):
                        chart_info = {
                            "sheet": sheet_name,
                            "index": i + 1,
                            "type": type(chart).__name__,
                            "title": getattr(chart, 'title', {}).get('text', f'Chart {i+1}') if hasattr(chart, 'title') else f'Chart {i+1}'
                        }
                        charts_info.append(chart_info)
            
            workbook.close()
            
            if charts_info:
                logger.info("Charts found in Excel", count=len(charts_info))
            
            return charts_info
            
        except Exception as e:
            logger.warning("Failed to extract charts from Excel", error=str(e))
            return []
    
    def format_output(self, content: Any, metadata: Dict[str, Any], 
                     output_format: str = "markdown") -> Dict[str, Any]:
        """重写格式化输出以处理Excel特定格式"""
        
        if output_format == "json":
            # 对于Excel，提供更结构化的JSON输出
            if isinstance(content, list) and all(isinstance(item, dict) and "sheet_name" in item for item in content):
                # 这是按工作表分页的内容
                structured_data = {
                    "sheets": content,
                    "metadata": metadata
                }
            else:
                structured_data = {
                    "content": content,
                    "metadata": metadata
                }
            
            return {
                "success": True,
                "data": structured_data,
                "format": output_format
            }
        
        # 对于其他格式，使用基类方法
        return super().format_output(content, metadata, output_format)
    
    async def _extract_images(self, excel_path: Path) -> List[Dict[str, Any]]:
        """从Excel文档中提取图片"""
        images_info = []
        
        try:
            # Excel文档也是一个ZIP文件
            with zipfile.ZipFile(str(excel_path), 'r') as excel_zip:
                # 查找图片文件
                image_files = [f for f in excel_zip.namelist() 
                             if f.startswith('xl/media/') and 
                             any(f.lower().endswith(ext) for ext in 
                                 ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'])]
                
                logger.info("Found images in Excel", count=len(image_files))
                
                for img_index, img_path in enumerate(image_files):
                    try:
                        # 读取图片数据
                        img_data = excel_zip.read(img_path)
                        
                        # 获取原始文件名和扩展名
                        original_filename = Path(img_path).name
                        
                        # 保存图片（使用实例ID避免文件名冲突）
                        filename = f"excel_{self.instance_id}_img_{img_index + 1}_{original_filename}"
                        image_info = await self.save_image(img_data, filename)
                        
                        # 添加额外信息
                        image_info.update({
                            "index": img_index + 1,
                            "original_path": img_path,
                            "original_filename": original_filename
                        })
                        
                        images_info.append(image_info)
                        
                        logger.info("Image extracted from Excel", 
                                   index=img_index + 1,
                                   filename=filename,
                                   size=len(img_data))
                        
                    except Exception as e:
                        logger.warning("Failed to extract image from Excel", 
                                     path=img_path,
                                     error=str(e))
                        continue
            
            logger.info("Image extraction from Excel completed", count=len(images_info))
            return images_info
            
        except Exception as e:
            logger.error("Failed to extract images from Excel document", error=str(e))
            return []
    
    async def _embed_image_urls(self, content: str, images_info: List[Dict[str, Any]]) -> str:
        """在Markdown内容中嵌入图片URL"""
        try:
            if not images_info:
                return content
            
            # 在Excel内容末尾添加图片引用部分
            image_section = ["\n\n## 📷 Extracted Images\n"]
            
            for i, img_info in enumerate(images_info, 1):
                filename = img_info.get("filename", f"image_{i}")
                url = img_info.get("url", "")
                size = img_info.get("size", 0)
                
                # 格式化文件大小
                if size > 1024 * 1024:
                    size_str = f"{size / (1024 * 1024):.1f} MB"
                elif size > 1024:
                    size_str = f"{size / 1024:.1f} KB"
                else:
                    size_str = f"{size} bytes"
                
                image_section.append(f"### Image {i}: {filename}")
                image_section.append(f"![{filename}]({url})")
                image_section.append(f"- **Size**: {size_str}")
                image_section.append(f"- **URL**: [{url}]({url})")
                image_section.append("")
            
            return content + "\n".join(image_section)
            
        except Exception as e:
            logger.warning("Failed to embed image URLs in Excel content", error=str(e))
            return content 